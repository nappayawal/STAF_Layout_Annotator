import json
import os
from dataclasses import dataclass
from typing import Dict, List, Tuple, Any

import openpyxl
from openpyxl.utils import range_boundaries, get_column_letter

import xlwings as xw


# -----------------------------
# Normalization
# -----------------------------
def normalize_tag(val: Any) -> str:
    """
    Normalize tag values for reliable matching between source and FLOOR PLAN.
    Rules:
      - cast to string
      - strip whitespace
      - uppercase
    """
    if val is None:
        return ""
    s = str(val).strip()
    return s.upper()


# -----------------------------
# Exceptions
# -----------------------------
class DuplicateTagError(Exception):
    pass


class TagColumnNotFoundError(Exception):
    pass


# -----------------------------
# Source reading (xlwings, supports .xls)
# -----------------------------
def read_source_table_xlwings(source_path: str) -> Tuple[List[str], List[List[Any]]]:
    """
    Read source table using xlwings so .xls works reliably.
    Returns: (headers, rows) where rows are list of row-values (same length as headers).
    """
    app = None
    wb = None
    try:
        app = xw.App(visible=False, add_book=False)
        app.display_alerts = False
        app.screen_updating = False

        wb = app.books.open(source_path)
        sht = wb.sheets[0]  # first sheet (like before .active)

        used = sht.used_range
        values = used.value  # 2D list or scalar

        if values is None:
            return [], []

        # Ensure 2D
        if not isinstance(values, list):
            return [], []
        if values and not isinstance(values[0], list):
            values = [values]

        # Remove fully empty trailing rows
        def row_has_any(r):
            return any(v is not None and str(v).strip() != "" for v in r)

        values = [r for r in values if row_has_any(r)]
        if not values:
            return [], []

        headers = [str(h).strip() if h is not None else "" for h in values[0]]
        rows = values[1:]

        return headers, rows
    finally:
        try:
            if wb is not None:
                wb.close()
        except Exception:
            pass
        try:
            if app is not None:
                app.quit()
        except Exception:
            pass


def build_comment_dict_from_source(
    source_path: str,
    log_callback=None
) -> Dict[str, str]:
    """
    Build { TAG -> formatted_comment_text } from Machine_Details.
    - Finds the 'Tag' column (case-insensitive).
    - Normalizes tags.
    - If duplicate tag is found: raise DuplicateTagError (stop program).
    """
    def log(msg: str):
        if log_callback:
            log_callback(msg)

    headers, rows = read_source_table_xlwings(source_path)

    if not headers:
        raise ValueError("Source file appears empty or unreadable.")

    # Find Tag column
    tag_col_idx = None
    for i, h in enumerate(headers):
        if normalize_tag(h) == "TAG":
            tag_col_idx = i
            break
    if tag_col_idx is None:
        raise TagColumnNotFoundError("Source file missing required column header: 'Tag'")

    comment_dict: Dict[str, str] = {}
    dup_tags: Dict[str, List[int]] = {}

    for r_index, row in enumerate(rows, start=2):  # +1 header row => excel row numbers
        # pad row length
        if len(row) < len(headers):
            row = row + [None] * (len(headers) - len(row))

        raw_tag = row[tag_col_idx] if tag_col_idx < len(row) else None
        tag = normalize_tag(raw_tag)

        if not tag:
            # Skip rows with blank tag
            continue

        if tag in comment_dict:
            dup_tags.setdefault(tag, []).append(r_index)
            continue

        #
        # Build comment text: Header: Value per column
        # - Skip Tag line entirely
        # - Force Position to show as integer text (251.0 -> 251)

        lines = []
        for h, v in zip(headers, row):
            h_clean = str(h).strip() if h is not None else ""
            if not h_clean:
                continue

            h_norm = normalize_tag(h_clean)  # reuse your normalize_tag() to compare headers safely

            # 1) Skip Tag in the comment output
            if h_norm == "TAG":
                continue

            # 2) Fix Position display: 251.0 -> 251
            if h_norm == "POSITION":
                if v is None or str(v).strip() == "":
                    v_str = ""
                else:
                    # if Excel gave float-like text/number, strip decimals safely
                    try:
                        v_str = str(int(float(v)))
                    except Exception:
                        # fallback: keep raw text but strip ".0" if present
                        s = str(v).strip()
                        v_str = s[:-2] if s.endswith(".0") else s
            else:
                v_str = "" if v is None else str(v).strip()

            lines.append(f"{h_clean}: {v_str}")

        comment_dict[tag] = "\n".join(lines)

    if dup_tags:
        # Also add the first occurrence row? We only tracked duplicates rows; include tag itself.
        msg_lines = ["Source file has duplicate Tag!"]
        for t, row_nums in dup_tags.items():
            msg_lines.append(f"  - {t} duplicates at rows: {', '.join(map(str, row_nums))}")
        raise DuplicateTagError("\n".join(msg_lines))

    log(f"✅ Source loaded. Unique tags: {len(comment_dict)}")
    return comment_dict


# -----------------------------
# Target scanning (openpyxl read-only; merged-aware)
# -----------------------------
def load_target_workbook_readonly(target_path: str):
    """
    Load STAF.xlsm with openpyxl for read-only scanning.
    We do NOT save with openpyxl.
    """
    # read_only=False so merged_cells are reliably available; we still won't save
    return openpyxl.load_workbook(target_path, keep_vba=True, data_only=True, read_only=False)


def build_merged_cell_lookup(ws) -> Dict[str, str]:
    """
    Map every cell coordinate in a merged range to the range's top-left coordinate.
    Example: 'B4' -> 'A4' if A4:B4 is merged (top-left A4)
    """
    lookup: Dict[str, str] = {}
    for merged in ws.merged_cells.ranges:
        min_col, min_row, max_col, max_row = merged.bounds
        top_left = f"{get_column_letter(min_col)}{min_row}"
        for r in range(min_row, max_row + 1):
            for c in range(min_col, max_col + 1):
                lookup[f"{get_column_letter(c)}{r}"] = top_left
    return lookup


def build_floor_tag_index(
    target_path: str,
    allowed_tags: set[str],
    sheet_name: str = "FLOOR PLAN",
    log_callback=None
) -> Dict[str, List[str]]:
    """
    Scan FLOOR PLAN once and build:
      { normalized_tag -> [top_left_cell_address] }

    IMPORTANT:
      - Only indexes cell values that match tags from the SOURCE file (allowed_tags).
      - Duplicate check is only for those allowed tags.
    """
    def log(msg: str):
        if log_callback:
            log_callback(msg)

    if not allowed_tags:
        raise ValueError("Allowed tag set is empty. Source may have no valid Tag values.")

    wb = load_target_workbook_readonly(target_path)
    if sheet_name not in wb.sheetnames:
        raise ValueError(f"Target file missing sheet: {sheet_name}")

    ws = wb[sheet_name]
    merged_lookup = build_merged_cell_lookup(ws)

    dim = ws.calculate_dimension()
    min_col, min_row, max_col, max_row = range_boundaries(dim)

    idx: Dict[str, List[str]] = {}
    dup_tags: Dict[str, List[str]] = {}

    scanned = 0
    matched = 0

    for row in ws.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
        for cell in row:
            scanned += 1
            v = cell.value
            if v is None:
                continue

            tag = normalize_tag(v)

            # ✅ FILTER: only treat it as a tag if it exists in source
            if tag not in allowed_tags:
                continue

            matched += 1
            coord = cell.coordinate
            top_left = merged_lookup.get(coord, coord)

            if tag in idx:
                idx[tag].append(top_left)
                dup_tags.setdefault(tag, []).append(top_left)
            else:
                idx[tag] = [top_left]

    log(f"✅ FLOOR PLAN scanned once. Cells scanned: {scanned}, matched source tags: {matched}")
    log(f"✅ Indexed allowed tags found on FLOOR PLAN: {len(idx)} (range {dim})")

    # Duplicate check ONLY for allowed tags found more than once
    if dup_tags:
        msg_lines = ["Target has a duplicate Tag!"]
        for t, _ in dup_tags.items():
            msg_lines.append(f"  - {t} found in cells: {', '.join(idx.get(t, []))}")
        raise DuplicateTagError("\n".join(msg_lines))

    return idx



# -----------------------------
# Placement planning + JSON reporting
# -----------------------------
@dataclass
class PlanResult:
    placements: List[Tuple[str, str]]  # (cell_address, note_text)
    missing_tags: List[str]
    total_source_tags: int
    total_target_tags: int


def plan_placements(
    comment_dict: Dict[str, str],
    floor_index: Dict[str, List[str]],
    log_callback=None
) -> PlanResult:
    """
    Create placements list by matching tags.
    If a source tag isn't found -> goes to missing_tags.
    """
    def log(msg: str):
        if log_callback:
            log_callback(msg)

    placements: List[Tuple[str, str]] = []
    missing: List[str] = []

    for tag, note_text in comment_dict.items():
        if tag not in floor_index:
            missing.append(tag)
            continue

        # Because duplicates in target are forbidden, we expect exactly one address
        for addr in floor_index[tag]:
            placements.append((addr, note_text))

    log(f"🧭 Placements planned: {len(placements)}")
    log(f"⚠ Missing tags (source not found on FLOOR PLAN): {len(missing)}")

    return PlanResult(
        placements=placements,
        missing_tags=missing,
        total_source_tags=len(comment_dict),
        total_target_tags=len(floor_index),
    )


def write_missing_tags_json(missing_tags: List[str], target_path: str, filename: str = "missing_tags.json") -> str:
    """
    Save missing tags JSON into the same folder as the STAF target file.
    Returns written path.
    """
    out_dir = os.path.dirname(os.path.abspath(target_path))
    out_path = os.path.join(out_dir, filename)
    payload = {
        "missing_count": len(missing_tags),
        "missing_tags": missing_tags,
    }
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2)
    return out_path
